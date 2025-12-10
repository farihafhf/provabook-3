from typing import Iterable
from datetime import datetime, date
import pytz

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

# ============================================================================
# COLOR DEFINITIONS FOR EXCEL EXPORT
# ============================================================================

# ETD colors (urgency-based)
ETD_URGENT_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
ETD_URGENT_FONT = Font(color="CC0000")  # Dark red
ETD_SOON_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
ETD_SOON_FONT = Font(color="B8860B")  # Dark amber

# Order Placement Date colors (recency-based)
RECENT_ORDER_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue
RECENT_ORDER_FONT = Font(color="0066CC")  # Dark blue
SOMEWHAT_RECENT_FILL = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")  # Very pale blue
SOMEWHAT_RECENT_FONT = Font(color="666666")  # Dark gray

# Submission/Approval colors
PENDING_SUBMISSION_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
PENDING_SUBMISSION_FONT = Font(color="B8860B")  # Dark amber
CLOSED_SUBMISSION_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue
CLOSED_SUBMISSION_FONT = Font(color="0066CC")  # Dark blue
APPROVED_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green
APPROVED_FONT = Font(color="006600")  # Dark green
REJECTED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
REJECTED_FONT = Font(color="CC0000")  # Dark red

# PP Sample special highlight (stronger orange for pending)
PP_SAMPLE_PENDING_FILL = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")  # Stronger orange
PP_SAMPLE_PENDING_FONT = Font(color="CC6600")  # Dark orange

# Bulk Start Date colors
BULK_ACTIVE_FILL = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")  # Light orange
BULK_ACTIVE_FONT = Font(color="CC6600")  # Dark orange

# Status-based colors
STATUS_COLORS = {
    'upcoming': (PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"), Font(color="666666")),
    'in_development': (PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"), Font(color="0066CC")),
    'running': (PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"), Font(color="B8860B")),
    'bulk': (PatternFill(start_color="E5CCFF", end_color="E5CCFF", fill_type="solid"), Font(color="660099")),
    'completed': (PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"), Font(color="006600")),
    'archived': (PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"), Font(color="666666")),
}

# ============================================================================

APPROVAL_HEADER_LABELS = {
    'aop': 'AOP Approval Date',
    'handloom': 'Handloom Approval Date',
    'labDip': 'Lab Dip Approval Date',
    'ppSample': 'PP Sample Approval Date',
    'price': 'Price Approval Date',
    'quality': 'Quality Approval Date',
}

SUBMISSION_HEADER_LABELS = {
    'aop': 'AOP Submission Date',
    'handloom': 'Handloom Submission Date',
    'labDip': 'Lab Dip Submission Date',
    'ppSample': 'PP Sample Submission Date',
    'price': 'Price Submission Date',
    'quality': 'Quality Submission Date',
}


def generate_orders_excel(queryset: Iterable, filters: dict = None) -> tuple:
    """
    Generate an Excel workbook for the given orders queryset.
    Returns (workbook, filename) tuple.
    
    Args:
        queryset: QuerySet of Order objects
        filters: Dictionary of applied filters for filename generation
    """
    from apps.orders.models_style_color import OrderStyle
    from apps.orders.models_order_line import OrderLine
    from apps.orders.models_supplier_delivery import SupplierDelivery
    from apps.orders.models_document import Document
    from apps.orders.models import ApprovalHistory
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders Export"
    
    # Get unique approval types across all orders for dynamic columns
    approval_types = set()
    for order in queryset:
        if order.approval_status:
            approval_types.update(order.approval_status.keys())
        # Also check approval history for all approval types
        for history in ApprovalHistory.objects.filter(order=order):
            approval_types.add(history.approval_type)
    
    # Sort approval types for consistent column order
    approval_types = sorted(list(approval_types))

    # Filter out approval types that should not appear as columns
    visible_approval_types = [
        t for t in approval_types
        if t not in {"bulkSwatch", "qualityTest", "strikeOff"}
    ]
    
    # Build header row - reordered with prices moved to end
    headers = [
        "Assigned To",
        "Buyer",
        "Order No.",
        "Style Number",
        "Description",
        "Color Code",
        "CAD Name",
        "Quantity Expected",
        "Unit",
        "ETD Date",
        "ETA Date",
        "ETD Quantity",
        "Order Placement Date",
        "Delivery Excess/Shortage",
    ]
    
    # Add dynamic approval stage columns (submission date before approval date)
    for approval_type in visible_approval_types:
        # Add submission date column first
        submission_label = SUBMISSION_HEADER_LABELS.get(
            approval_type,
            f"approval_stage_{approval_type}_submission_date",
        )
        headers.append(submission_label)
        
        # Then add approval date column
        header_label = APPROVAL_HEADER_LABELS.get(
            approval_type,
            f"approval_stage_{approval_type}_date",
        )
        headers.append(header_label)
    
    headers.extend([
        "Order Status",
        "Bulk Start Date",
        "Latest PI Sent Date",
        "LC Received Date",
        "Mill Price",
        "Prova Price",
        "Profit per Unit",
        "Total Commission",
        "Adjusted Profit",
        "Reduced Profit",
        "Delivery History",
    ])
    
    worksheet.append(headers)
    
    # Format header row
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Freeze header row and first 3 columns (Assigned To, Buyer, Order No.)
    worksheet.freeze_panes = 'D2'
    
    # Process each order
    dhaka_tz = pytz.timezone('Asia/Dhaka')
    
    # Extract status filter if provided - used to filter lines during export
    status_filter = filters.get('status') if filters else None
    
    for order in queryset:
        # Get all styles for this order
        styles = OrderStyle.objects.filter(order=order).prefetch_related('lines')
        
        if not styles.exists():
            # No styles - export order-level data only (skip if status filter is applied)
            if status_filter:
                continue
            row_data = _build_order_row(
                order, None, None, visible_approval_types, dhaka_tz
            )
            worksheet.append(row_data)
        else:
            # Export each order line
            for style in styles:
                # Filter lines by status if status filter is applied
                if status_filter:
                    lines = style.lines.filter(status=status_filter)
                else:
                    lines = style.lines.all()
                
                if not lines.exists():
                    # Style has no lines matching filter - skip or export style-level data
                    # If status filter is applied, skip styles without matching lines
                    if status_filter:
                        continue
                    row_data = _build_order_row(
                        order, style, None, visible_approval_types, dhaka_tz
                    )
                    worksheet.append(row_data)
                else:
                    # Export each line
                    for line in lines:
                        row_data = _build_order_row(
                            order, style, line, visible_approval_types, dhaka_tz
                        )
                        worksheet.append(row_data)
    
    # Apply color coding to the worksheet
    _apply_color_coding(worksheet, headers, visible_approval_types, dhaka_tz)
    
    # Auto-size columns
    for idx, column in enumerate(worksheet.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(idx)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename
    filename = _generate_filename(filters, dhaka_tz)
    
    return workbook, filename


def _build_order_row(order, style, line, approval_types, dhaka_tz):
    """Build a single row of data for Excel export."""
    from apps.orders.models_supplier_delivery import SupplierDelivery
    from apps.orders.models_document import Document
    from apps.orders.models import ApprovalHistory
    
    # Helper function to format datetime
    def format_datetime(dt):
        if dt is None:
            return ""
        if isinstance(dt, datetime):
            if dt.tzinfo is None:
                dt = pytz.utc.localize(dt)
            dt = dt.astimezone(dhaka_tz)
            return dt.strftime("%Y-%m-%d %I:%M %p")
        else:
            # Date only
            return dt.strftime("%Y-%m-%d")
    
    # Determine data source (line > style > order)
    data_source = line if line else (style if style else order)
    
    # ID
    id_value = order.order_number
    
    # Order No.
    order_no = order.order_number
    
    # Style Number
    style_number = style.style_number if style else (order.style_number or "")
    
    # Color
    color = ""
    if line and line.color_code:
        color = f"{line.color_code}"
        if line.color_name:
            color += f" ({line.color_name})"
    
    # CAD
    cad = ""
    if line and line.cad_code:
        cad = f"{line.cad_code}"
        if line.cad_name:
            cad += f" ({line.cad_name})"
    elif order.cad:
        cad = order.cad
    
    # Description
    description_parts = []
    fabric_type = getattr(style, 'fabric_type', None) or order.fabric_type
    if fabric_type:
        description_parts.append(f"**Fabric Type** {fabric_type}")
    
    composition = getattr(style, 'fabric_composition', None) or order.fabric_composition
    if composition:
        description_parts.append(f"Composition: {composition}")
    
    gsm = getattr(style, 'gsm', None) or order.gsm
    if gsm:
        description_parts.append(f"GSM: {gsm}")
    
    finish_type = getattr(style, 'finish_type', None) or order.finish_type
    if finish_type:
        description_parts.append(f"Finish Type: {finish_type}")
    
    construction = getattr(style, 'construction', None) or order.construction
    if construction:
        description_parts.append(f"Construction: {construction}")
    
    description = "\n".join(description_parts)
    
    # Buyer
    buyer = order.buyer_name or order.customer_name
    
    # Supplier
    supplier = getattr(data_source, 'mill_name', None) or order.mill_name or ""
    
    # Assigned To
    assigned_to = order.merchandiser.full_name if order.merchandiser else ""
    
    # Quantity Expected
    quantity_expected = float(data_source.quantity) if hasattr(data_source, 'quantity') and data_source.quantity else 0
    
    # Unit
    unit = getattr(data_source, 'unit', 'meters')
    
    # Mill Price
    mill_price = float(data_source.mill_price) if hasattr(data_source, 'mill_price') and data_source.mill_price else 0
    
    # LC Open Price (Prova Price)
    lc_open_price = float(data_source.prova_price) if hasattr(data_source, 'prova_price') and data_source.prova_price else 0
    
    # Profit per Unit
    profit_per_unit = lc_open_price - mill_price
    
    # Original Profit
    original_profit = profit_per_unit * quantity_expected
    
    # ETD Date
    etd_date = format_datetime(getattr(data_source, 'etd', None) or order.etd)
    
    # ETA Date
    eta_date = format_datetime(getattr(data_source, 'eta', None) or order.eta)
    
    # ETD Quantity & ETD Total Quantity
    # Get deliveries for this order
    deliveries = SupplierDelivery.objects.filter(order=order)
    if style:
        deliveries = deliveries.filter(style=style)
    
    etd_quantity = sum(float(d.delivered_quantity) for d in deliveries)
    etd_total_quantity = float(order.total_delivered_quantity)
    
    # Order Placement Date
    order_placement_date = format_datetime(order.order_date) if order.order_date else ""
    
    # Delivery Excess/Shortage
    delivery_excess_shortage = etd_quantity - quantity_expected
    
    # Approval stage dates (dynamic columns - submission date before approval date)
    approval_data = []  # Will contain pairs of (submission_date, approval_date)
    for approval_type in approval_types:
        # Get the submission date (first 'submission' status)
        submission_date = ""
        
        # Check line-level submission first
        if line:
            submission_history = ApprovalHistory.objects.filter(
                order=order,
                order_line=line,
                approval_type=approval_type,
                status='submission'
            ).order_by('created_at').first()
            
            if submission_history:
                submission_date = format_datetime(submission_history.created_at)
        
        # Fallback to order-level submission
        if not submission_date:
            submission_history = ApprovalHistory.objects.filter(
                order=order,
                order_line__isnull=True,
                approval_type=approval_type,
                status='submission'
            ).order_by('created_at').first()
            
            if submission_history:
                submission_date = format_datetime(submission_history.created_at)
        
        # Get the latest approved date for this approval type
        approval_date = ""
        
        # Check line-level approval history first
        if line:
            history = ApprovalHistory.objects.filter(
                order=order,
                order_line=line,
                approval_type=approval_type,
                status='approved'
            ).order_by('-created_at').first()
            
            if history:
                approval_date = format_datetime(history.created_at)
        
        # Fallback to order-level approval history
        if not approval_date:
            history = ApprovalHistory.objects.filter(
                order=order,
                order_line__isnull=True,
                approval_type=approval_type,
                status='approved'
            ).order_by('-created_at').first()
            
            if history:
                approval_date = format_datetime(history.created_at)
        
        # Add both submission and approval dates
        approval_data.append(submission_date)
        approval_data.append(approval_date)
    
    # Order Status (use order-level status)
    order_status = order.get_status_display()
    
    # Bulk Start Date
    bulk_start_date = ""
    if order.status == 'bulk':
        # Try to find when status changed to bulk
        # This would require status change tracking - for now use updated_at
        bulk_start_date = format_datetime(order.updated_at)
    
    # Latest PI Sent Date
    pi_sent_date = ""
    pi_docs = Document.objects.filter(order=order, category='pi').order_by('-created_at').first()
    if pi_docs:
        pi_sent_date = format_datetime(pi_docs.created_at)
    
    # LC Received Date
    lc_received_date = ""
    lc_docs = Document.objects.filter(order=order, category='lc').order_by('-created_at').first()
    if lc_docs:
        lc_received_date = format_datetime(lc_docs.created_at)
    
    # Total Commission
    commission = getattr(data_source, 'commission', None)
    total_commission = float(commission) * quantity_expected if commission else 0
    
    # Adjusted Profit (based on delivered quantity)
    if etd_total_quantity > 0:
        delivery_ratio = min(1.0, etd_total_quantity / float(order.quantity))
    else:
        delivery_ratio = 0.0
    adjusted_profit = original_profit * delivery_ratio
    
    # Reduced Profit
    reduced_profit = adjusted_profit - original_profit
    
    # Supplier Delivery History
    delivery_history_parts = []
    for delivery in deliveries.order_by('delivery_date'):
        delivery_history_parts.append(
            f"{delivery.delivery_date.strftime('%Y-%m-%d')}:{delivery.delivered_quantity}"
        )
    supplier_delivery_history = "; ".join(delivery_history_parts)
    
    # Build row - reordered to match new header layout
    row = [
        assigned_to,
        buyer,
        order_no,
        style_number,
        description,
        color,
        cad,
        quantity_expected,
        unit,
        etd_date,
        eta_date,
        etd_quantity,
        order_placement_date,
        delivery_excess_shortage,
    ]
    
    # Add approval data (submission date + approval date pairs)
    row.extend(approval_data)
    
    # Add remaining columns with prices moved before Total Commission
    row.extend([
        order_status,
        bulk_start_date,
        pi_sent_date,
        lc_received_date,
        mill_price,
        lc_open_price,
        profit_per_unit,
        total_commission,
        adjusted_profit,
        reduced_profit,
        supplier_delivery_history,
    ])
    
    return row


def _generate_filename(filters, dhaka_tz):
    """Generate filename based on filters and current timestamp."""
    # Build filter description
    filter_parts = []
    
    if filters:
        if filters.get('status'):
            filter_parts.append(filters['status'])
        if filters.get('category'):
            filter_parts.append(filters['category'])
        if filters.get('merchandiser'):
            filter_parts.append('merchandiser')
        if filters.get('search'):
            filter_parts.append('search')
    
    filter_desc = "_".join(filter_parts) if filter_parts else "all"
    
    # Get current timestamp in Asia/Dhaka timezone
    now = datetime.now(dhaka_tz)
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%I%M%p")
    
    filename = f"prova_orders_export_{filter_desc}_{date_str}_{time_str}.xlsx"
    
    return filename


def _apply_color_coding(worksheet, headers, approval_types, dhaka_tz):
    """
    Apply color coding to the worksheet based on business rules.
    
    Colors applied:
    - ETD Date: urgency-based (red for overdue/urgent, yellow for soon)
    - Order Placement Date: recency-based (blue for recent)
    - Submission/Approval pairs: pending/approved status colors
    - Bulk Start Date: orange when present
    - Order Status: status-specific colors
    """
    # Get today's date in Dhaka timezone
    today = datetime.now(dhaka_tz).date()
    
    # Helper function to parse date from cell value
    def parse_date_from_cell(cell_value):
        if not cell_value or cell_value == "":
            return None
        try:
            # Format is "YYYY-MM-DD HH:MM AM/PM" or "YYYY-MM-DD"
            date_str = str(cell_value).split(" ")[0]
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except (ValueError, IndexError):
            return None
    
    # Get column indices (1-indexed for openpyxl)
    col_indices = {}
    for idx, header in enumerate(headers, start=1):
        col_indices[header] = idx
    
    # Get specific column indices
    etd_col = col_indices.get("ETD Date")
    order_placement_col = col_indices.get("Order Placement Date")
    bulk_start_col = col_indices.get("Bulk Start Date")
    status_col = col_indices.get("Order Status")
    
    # Get submission/approval column pairs
    approval_col_pairs = {}
    for approval_type in approval_types:
        submission_header = SUBMISSION_HEADER_LABELS.get(
            approval_type,
            f"approval_stage_{approval_type}_submission_date"
        )
        approval_header = APPROVAL_HEADER_LABELS.get(
            approval_type,
            f"approval_stage_{approval_type}_date"
        )
        
        sub_col = col_indices.get(submission_header)
        app_col = col_indices.get(approval_header)
        
        if sub_col and app_col:
            approval_col_pairs[approval_type] = (sub_col, app_col)
    
    # Status value to key mapping (display value -> internal key)
    status_display_to_key = {
        'Upcoming': 'upcoming',
        'In Development': 'in_development',
        'Running Order': 'running',
        'Bulk': 'bulk',
        'Completed': 'completed',
        'Archived': 'archived',
    }
    
    # Iterate through data rows (skip header row)
    for row_idx in range(2, worksheet.max_row + 1):
        # -----------------------------------------------------------------
        # 1. ETD Date coloring (urgency-based)
        # -----------------------------------------------------------------
        if etd_col:
            etd_cell = worksheet.cell(row=row_idx, column=etd_col)
            etd_date = parse_date_from_cell(etd_cell.value)
            
            if etd_date:
                diff_days = (etd_date - today).days
                
                if diff_days <= 5:  # Past or within 5 days - urgent
                    etd_cell.fill = ETD_URGENT_FILL
                    etd_cell.font = ETD_URGENT_FONT
                elif 6 <= diff_days <= 10:  # 6-10 days - upcoming soon
                    etd_cell.fill = ETD_SOON_FILL
                    etd_cell.font = ETD_SOON_FONT
                # else: no special fill
        
        # -----------------------------------------------------------------
        # 2. Order Placement Date coloring (recency-based)
        # -----------------------------------------------------------------
        if order_placement_col:
            placement_cell = worksheet.cell(row=row_idx, column=order_placement_col)
            placement_date = parse_date_from_cell(placement_cell.value)
            
            if placement_date:
                days_ago = (today - placement_date).days
                
                if days_ago <= 7:  # Last 7 days - very recent
                    placement_cell.fill = RECENT_ORDER_FILL
                    placement_cell.font = RECENT_ORDER_FONT
                elif 8 <= days_ago <= 30:  # 8-30 days - recent but not new
                    placement_cell.fill = SOMEWHAT_RECENT_FILL
                    placement_cell.font = SOMEWHAT_RECENT_FONT
                # else: no fill for older
        
        # -----------------------------------------------------------------
        # 3. Submission/Approval Date pair coloring
        # -----------------------------------------------------------------
        for approval_type, (sub_col, app_col) in approval_col_pairs.items():
            sub_cell = worksheet.cell(row=row_idx, column=sub_col)
            app_cell = worksheet.cell(row=row_idx, column=app_col)
            
            has_submission = sub_cell.value and str(sub_cell.value).strip() != ""
            has_approval = app_cell.value and str(app_cell.value).strip() != ""
            
            if has_submission and has_approval:
                # Both present - submission closed, approved
                sub_cell.fill = CLOSED_SUBMISSION_FILL
                sub_cell.font = CLOSED_SUBMISSION_FONT
                app_cell.fill = APPROVED_FILL
                app_cell.font = APPROVED_FONT
            elif has_submission and not has_approval:
                # Submission pending approval
                if approval_type == 'ppSample':
                    # Special highlight for PP Sample pending
                    sub_cell.fill = PP_SAMPLE_PENDING_FILL
                    sub_cell.font = PP_SAMPLE_PENDING_FONT
                else:
                    sub_cell.fill = PENDING_SUBMISSION_FILL
                    sub_cell.font = PENDING_SUBMISSION_FONT
            # else: neither present - no fill
        
        # -----------------------------------------------------------------
        # 4. Bulk Start Date coloring
        # -----------------------------------------------------------------
        if bulk_start_col:
            bulk_cell = worksheet.cell(row=row_idx, column=bulk_start_col)
            
            if bulk_cell.value and str(bulk_cell.value).strip() != "":
                bulk_cell.fill = BULK_ACTIVE_FILL
                bulk_cell.font = BULK_ACTIVE_FONT
        
        # -----------------------------------------------------------------
        # 5. Order Status coloring
        # -----------------------------------------------------------------
        if status_col:
            status_cell = worksheet.cell(row=row_idx, column=status_col)
            status_display = str(status_cell.value) if status_cell.value else ""
            
            # Map display value to internal key
            status_key = status_display_to_key.get(status_display)
            
            if status_key and status_key in STATUS_COLORS:
                fill, font = STATUS_COLORS[status_key]
                status_cell.fill = fill
                status_cell.font = font


def generate_purchase_order_pdf(order, buffer):
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 25 * mm
    y = height - margin

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(margin, y, "Provabook ERP")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(margin, y - 14, "Company Name / Logo")

    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawCentredString(width / 2, y - 40, "PURCHASE ORDER")

    details_y = y - 80
    line_height = 14

    def draw_label_value(label, value, x_label, x_value, y_pos):
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(x_label, y_pos, f"{label}:")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(x_value, y_pos, value if value is not None else "-")

    right_x_label = width - margin - 130
    right_x_value = width - margin - 60

    order_identifier = getattr(order, "order_number", None) or str(getattr(order, "id", ""))
    draw_label_value("Order ID", order_identifier, right_x_label, right_x_value, details_y)
    draw_label_value("Customer", getattr(order, "customer_name", None), right_x_label, right_x_value, details_y - line_height)
    draw_label_value("Style No.", getattr(order, "style_number", None), right_x_label, right_x_value, details_y - 2 * line_height)

    etd_value = getattr(order, "etd", None)
    etd_str = etd_value.strftime("%Y-%m-%d") if etd_value else None
    draw_label_value("ETD", etd_str, right_x_label, right_x_value, details_y - 3 * line_height)

    table_top = details_y - 60
    row_height = 18
    col_x = [margin, margin + 200, margin + 300, width - margin]

    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(col_x[0], table_top - row_height, col_x[-1] - col_x[0], row_height, stroke=0, fill=1)

    pdf.setFillColorRGB(0, 0, 0)
    pdf.setFont("Helvetica-Bold", 10)
    header_y = table_top - row_height + 5
    pdf.drawString(col_x[0] + 4, header_y, "Item")
    pdf.drawString(col_x[1] + 4, header_y, "Quantity")
    pdf.drawString(col_x[2] + 4, header_y, "Unit Price")
    pdf.drawString(col_x[3] - 70, header_y, "Total")

    data_y = table_top - 2 * row_height + 4
    pdf.setFont("Helvetica", 10)

    description = "Fabric order"
    style_number = getattr(order, "style_number", None)
    if style_number:
        description = f"Fabric order {style_number}"

    quantity_value = getattr(order, "quantity", None)
    unit_value = getattr(order, "unit", "")
    quantity_text = f"{quantity_value} {unit_value}".strip() if quantity_value is not None else ""

    currency = getattr(order, "currency", "")
    prova_price = getattr(order, "prova_price", None)
    mill_price = getattr(order, "mill_price", None)
    unit_price_value = prova_price if prova_price is not None else mill_price
    unit_price_text = f"{unit_price_value} {currency}".strip() if unit_price_value is not None else ""

    total_text = ""
    if unit_price_value is not None and quantity_value is not None:
        try:
            total_value = float(unit_price_value) * float(quantity_value)
            total_text = f"{total_value:.2f} {currency}".strip()
        except (TypeError, ValueError):
            total_text = ""

    pdf.drawString(col_x[0] + 4, data_y, description)
    if quantity_text:
        pdf.drawString(col_x[1] + 4, data_y, quantity_text)
    if unit_price_text:
        pdf.drawString(col_x[2] + 4, data_y, unit_price_text)
    if total_text:
        pdf.drawString(col_x[3] - 70, data_y, total_text)

    for x in col_x:
        pdf.line(x, table_top, x, table_top - 3 * row_height)

    pdf.line(col_x[0], table_top, col_x[-1], table_top)
    pdf.line(col_x[0], table_top - row_height, col_x[-1], table_top - row_height)
    pdf.line(col_x[0], table_top - 2 * row_height, col_x[-1], table_top - 2 * row_height)
    pdf.line(col_x[0], table_top - 3 * row_height, col_x[-1], table_top - 3 * row_height)

    footer_y = margin / 2
    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawCentredString(width / 2, footer_y, "Generated by Provabook")

    pdf.showPage()
    pdf.save()


def generate_tna_excel(queryset: Iterable, filters: dict = None) -> tuple:
    """
    Generate a TnA (Time and Action) Excel workbook for local orders.
    Format matches TNA-Tubulor.xls template.
    
    Returns (workbook, filename) tuple.
    
    Args:
        queryset: QuerySet of Order objects (filtered for local orders)
        filters: Dictionary of applied filters for filename generation
    """
    from apps.orders.models_style_color import OrderStyle
    from apps.orders.models_order_line import OrderLine
    from apps.orders.models import ApprovalHistory
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TnA Export"
    
    dhaka_tz = pytz.timezone('Asia/Dhaka')
    today = datetime.now(dhaka_tz)
    
    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Helper function to format date
    def format_date(dt):
        if dt is None:
            return ""
        if isinstance(dt, datetime):
            return dt.strftime("%Y-%m-%d")
        return dt.strftime("%Y-%m-%d") if dt else ""
    
    # Get unique buyer names from the orders
    buyer_names = set()
    for order in queryset:
        buyer_name = order.buyer_name or order.customer_name
        if buyer_name:
            buyer_names.add(buyer_name)
    
    buyer_display = ", ".join(sorted(buyer_names)) if buyer_names else "All Buyers"
    
    # Row 1: Company name
    worksheet['A1'] = "Prova Fashion & Accessories"
    worksheet['A1'].font = title_font
    worksheet.merge_cells('A1:E1')
    
    # Row 2: Buyer
    worksheet['A2'] = f"Buyer: {buyer_display}"
    worksheet['A2'].font = Font(bold=True)
    worksheet.merge_cells('A2:E2')
    
    # Row 3: Date
    worksheet['A3'] = f"DATE :"
    worksheet['A3'].font = Font(bold=True)
    worksheet['B3'] = today.strftime("%Y-%m-%d")
    
    # Row 4: Empty
    # Row 5: Headers
    headers = [
        "STYLE NAME",           # A
        "FABRICS",              # B
        "COLOR",                # C
        "ORDER QTY",            # D
        "YARN\nREQ",            # E
        "YARN\nBOOKED",         # F
        "YARN\nRCVD",           # G
        "LABDIP\nSUB",          # H
        "LABDIP\nAPPROVED",     # I
        "PP Yds",               # J
        "FIT CUM PP \nSUBMITE", # K
        "FIT CUM PP \nCOMMENTS",# L
        "KNITTING\nSTART",      # M
        "KNITTING\nCOMPLTE",    # N
        "DYEING\nSTART",        # O
        "DYEING\nCOMPLTE",      # P
        "1st BULK SUBMISSION",  # Q
        "LINE",                 # R
        "BULK SIZE SET",        # S
        "CUTTING\nSTART",       # T
        "CUTTING\nCOMPLTE",     # U
        "PRINT\nSEND",          # V
        "PRINT\nRCVD",          # W
        "SEWING\nINPUT",        # X
        "SEWING\nFINISH",       # Y
        "PACKING\nCOMPLTE",     # Z
        "FINAL\nINSPECTION",    # AA
        "EX-FACTORY",           # AB
        "REMARKS",              # AC
    ]
    
    header_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Process orders and lines
    data_row = 6
    total_quantity = 0
    line_counter = 0
    
    # Extract status filter if provided - used to filter lines during export
    status_filter = filters.get('status') if filters else None
    
    for order in queryset:
        # Get all styles for this order
        styles = OrderStyle.objects.filter(order=order).prefetch_related('lines')
        
        for style in styles:
            # Filter lines by status if status filter is applied
            if status_filter:
                lines = style.lines.filter(status=status_filter)
            else:
                lines = style.lines.all()
            
            if not lines.exists():
                # Style has no lines matching filter - skip
                continue
            
            for line in lines:
                line_counter += 1
                
                # Get fabric info (style-level or order-level)
                fabric_parts = []
                fabric_composition = style.fabric_composition or order.fabric_composition
                gsm = style.gsm or order.gsm
                if fabric_composition:
                    fabric_parts.append(fabric_composition)
                if gsm:
                    fabric_parts.append(f"{gsm} GSM")
                fabric_info = " ".join(fabric_parts) if fabric_parts else (order.fabric_type or "")
                
                # Get labdip dates from approval history
                labdip_sub_date = ""
                labdip_approved_date = ""
                
                # Check line-level approval history first
                labdip_sub_history = ApprovalHistory.objects.filter(
                    order=order,
                    order_line=line,
                    approval_type__in=['labDip', 'lab_dip'],
                    status='submission'
                ).order_by('created_at').first()
                
                if labdip_sub_history:
                    labdip_sub_date = format_date(labdip_sub_history.created_at)
                
                labdip_approved_history = ApprovalHistory.objects.filter(
                    order=order,
                    order_line=line,
                    approval_type__in=['labDip', 'lab_dip'],
                    status='approved'
                ).order_by('-created_at').first()
                
                if labdip_approved_history:
                    labdip_approved_date = format_date(labdip_approved_history.created_at)
                
                # Fallback to order-level if no line-level history
                if not labdip_sub_date:
                    order_labdip_sub = ApprovalHistory.objects.filter(
                        order=order,
                        order_line__isnull=True,
                        approval_type__in=['labDip', 'lab_dip'],
                        status='submission'
                    ).order_by('created_at').first()
                    if order_labdip_sub:
                        labdip_sub_date = format_date(order_labdip_sub.created_at)
                
                if not labdip_approved_date:
                    order_labdip_approved = ApprovalHistory.objects.filter(
                        order=order,
                        order_line__isnull=True,
                        approval_type__in=['labDip', 'lab_dip'],
                        status='approved'
                    ).order_by('-created_at').first()
                    if order_labdip_approved:
                        labdip_approved_date = format_date(order_labdip_approved.created_at)
                
                # Get bulk submission date
                bulk_submission_date = ""
                bulk_history = ApprovalHistory.objects.filter(
                    order=order,
                    approval_type__in=['bulkSwatch', 'bulk_swatch'],
                    status='submission'
                ).order_by('created_at').first()
                if bulk_history:
                    bulk_submission_date = format_date(bulk_history.created_at)
                
                # Build row data
                quantity = float(line.quantity) if line.quantity else 0
                total_quantity += quantity
                
                row_data = [
                    style.style_number or "",                           # STYLE NAME
                    fabric_info,                                        # FABRICS
                    line.color_code or "",                              # COLOR
                    quantity if quantity > 0 else "",                   # ORDER QTY
                    float(line.yarn_required) if line.yarn_required else "",  # YARN REQ
                    format_date(line.yarn_booked_date),                 # YARN BOOKED
                    format_date(line.yarn_received_date),               # YARN RCVD
                    labdip_sub_date,                                    # LABDIP SUB
                    labdip_approved_date,                               # LABDIP APPROVED
                    float(line.pp_yards) if line.pp_yards else "",      # PP Yds
                    format_date(line.fit_cum_pp_submit_date),           # FIT CUM PP SUBMITE
                    format_date(line.fit_cum_pp_comments_date),         # FIT CUM PP COMMENTS
                    format_date(line.knitting_start_date),              # KNITTING START
                    format_date(line.knitting_complete_date),           # KNITTING COMPLTE
                    format_date(line.dyeing_start_date),                # DYEING START
                    format_date(line.dyeing_complete_date),             # DYEING COMPLTE
                    bulk_submission_date,                               # 1st BULK SUBMISSION
                    chr(64 + (line_counter % 26) + 1) if line_counter <= 26 else str(line_counter),  # LINE (A, B, C...)
                    format_date(line.bulk_size_set_date),               # BULK SIZE SET
                    format_date(line.cutting_start_date),               # CUTTING START
                    format_date(line.cutting_complete_date),            # CUTTING COMPLTE
                    format_date(line.print_send_date),                  # PRINT SEND
                    format_date(line.print_received_date),              # PRINT RCVD
                    format_date(line.sewing_input_date),                # SEWING INPUT
                    format_date(line.sewing_finish_date),               # SEWING FINISH
                    format_date(line.packing_complete_date),            # PACKING COMPLTE
                    format_date(line.final_inspection_date),            # FINAL INSPECTION
                    format_date(line.ex_factory_date),                  # EX-FACTORY
                    line.notes or "",                                   # REMARKS
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=data_row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                data_row += 1
    
    # Total row
    total_row = data_row
    worksheet.cell(row=total_row, column=1, value="TOTAL QYUANTITY =")
    worksheet.cell(row=total_row, column=1).font = Font(bold=True)
    worksheet.cell(row=total_row, column=1).fill = total_fill
    worksheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    
    worksheet.cell(row=total_row, column=4, value=total_quantity)
    worksheet.cell(row=total_row, column=4).font = Font(bold=True)
    worksheet.cell(row=total_row, column=4).fill = total_fill
    
    # Apply borders to total row
    for col_idx in range(1, len(headers) + 1):
        cell = worksheet.cell(row=total_row, column=col_idx)
        cell.border = thin_border
    
    # Auto-size columns
    for idx, column in enumerate(worksheet.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(idx)
        
        for cell in column:
            try:
                if cell.value:
                    cell_value = str(cell.value)
                    # Handle multi-line headers
                    lines = cell_value.split('\n')
                    max_line_length = max(len(line) for line in lines)
                    max_length = max(max_length, max_line_length)
            except:
                pass
        
        # Set minimum width and cap at 20
        adjusted_width = max(min(max_length + 2, 20), 10)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Set row height for header row to accommodate multi-line text
    worksheet.row_dimensions[header_row].height = 35
    
    # Generate filename
    filename = _generate_tna_filename(filters, dhaka_tz)
    
    return workbook, filename


def _generate_tna_filename(filters, dhaka_tz):
    """Generate filename for TnA export based on current timestamp."""
    now = datetime.now(dhaka_tz)
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%I%M%p")
    
    filename = f"TnA_Export_{date_str}_{time_str}.xlsx"
    
    return filename
